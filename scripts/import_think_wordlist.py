#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Import Cambridge <Think> textbook bilingual wordlists into the dev DB.

Pipeline:
  parse 5 Excels  ->  translate EN Definition -> ZH via DeepSeek-V3
                 ->  idempotently INSERT into word_bank / level / word_list / word_list_item
                 ->  evict Redis `word:levels` cache

DISCIPLINE (iron rule of this project):
  * INSERT-only. Never DROP / TRUNCATE / ALTER / DELETE on existing rows.
  * Never touch user / study / progress / wrong-word tables.
  * Idempotent: safe to re-run. Re-runs skip translation via cache.

Usage (from repo root):
  python scripts/import_think_wordlist.py            # use cached translations only
  python scripts/import_think_wordlist.py --translate # allow DeepSeek calls (costs ~RMB 0.1)
"""

import argparse
import glob
import json
import os
import re
import sys
import time
import unicodedata

import openpyxl
import pymysql

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

WORDLIST_DIR = "wordlist"
CACHE_FILE = os.path.join(WORDLIST_DIR, "_zh_cache.json")

# filename-fragment -> (level_code, name_zh, name_en, sort_order, target_word_count)
LEVEL_MAP = [
    ("Think_Starter", "THINK_STARTER", "Think 入门", "Think Starter", 100, 596),
    ("Think_Level_2", "THINK_L2",      "Think 2",     "Think Level 2",  101, 489),
    ("Think_Level_3", "THINK_L3",      "Think 3",     "Think Level 3",  102, 470),
    ("Think_Level_4", "THINK_L4",      "Think 4",     "Think Level 4",  103, 566),
    ("Think_Level_5", "THINK_L5",      "Think 5",     "Think Level 5",  104, 518),
]

# DB connection (local docker). Port verified via `docker ps`.
DB_CONFIG = {
    "host": "127.0.0.1",
    "port": 3306,
    "user": "root",
    "password": "root",
    "database": "vocabmaster",
    "charset": "utf8mb4",
    "autocommit": False,
}

# DeepSeek
DEEPSEEK_URL = "https://api.deepseek.com/chat/completions"
DEEPSEEK_MODEL = "deepseek-chat"
DEEPSEEK_KEY = os.environ.get("DEEPSEEK_API_KEY")  # 对齐项目约定（generate_examples.py 同名），密钥勿硬编码
BATCH_SIZE = 30
SLEEP_BETWEEN_BATCHES = 0.3

# Redis evict (via docker exec to keep deps light)
REDIS_DEL_CMD = [
    "docker", "exec", "vocab-redis",
    "redis-cli", "-a", "redis_prod_2024", "DEL", "word:levels",
]
REDIS_PATTERN_DEL_CMD = [
    "docker", "exec", "vocab-redis",
    "redis-cli", "-a", "redis_prod_2024",
    "--scan", "--pattern", "word:levels*", "DEL",
]


# ---------------------------------------------------------------------------
# Excel parsing
# ---------------------------------------------------------------------------

def match_level_for_file(filename):
    fn_lower = os.path.basename(filename).lower()
    for frag, code, zh, en, sort, target in LEVEL_MAP:
        if frag.lower() in fn_lower:
            return frag, code, zh, en, sort, target
    return None


def parse_excel(path):
    """Return list of dicts: word, unit_no, page, en_def, pos, example, ipa."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = []
    for idx, r in enumerate(ws.iter_rows(values_only=True)):
        if idx == 0:
            # header row — skip. (verified all 5 files have a header)
            continue
        word = r[0]
        if word is None or str(word).strip() == "":
            continue
        rows.append({
            "word": str(word).strip(),
            "unit_no": int(r[1]) if r[1] is not None else 0,
            "page": int(r[2]) if r[2] is not None else None,
            "en_def": (str(r[3]).strip() if r[3] is not None else "") or "",
            "pos": (str(r[4]).strip() if r[4] is not None else None),
            "example": (str(r[5]).strip() if r[5] is not None else None),
            "ipa": (str(r[7]).strip() if r[7] is not None else None),
        })
    wb.close()
    return rows


# ---------------------------------------------------------------------------
# Translation (EN Definition -> ZH) via DeepSeek
# ---------------------------------------------------------------------------

def load_cache():
    if os.path.exists(CACHE_FILE):
        try:
            with open(CACHE_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return {}
    return {}


def save_cache(cache):
    with open(CACHE_FILE, "w", encoding="utf-8") as f:
        json.dump(cache, f, ensure_ascii=False, indent=0)


def call_deepseek(batch):
    """batch: list of (word, en_def). Returns dict word->zh or None on error."""
    import urllib.request
    import urllib.error

    user_payload = json.dumps(
        [{"w": w, "def": d} for w, d in batch],
        ensure_ascii=False,
    )
    sys_msg = (
        "你是英汉词典翻译。把每个英文定义翻译成简明中文释义（不超过该英文词数×3个汉字）。"
        "输出纯 JSON 数组，元素形如 {\"w\":原词,\"zh\":中文释义}，"
        "不要输出任何解释、代码块标记或多余文字。"
    )
    body = json.dumps({
        "model": DEEPSEEK_MODEL,
        "messages": [
            {"role": "system", "content": sys_msg},
            {"role": "user", "content": user_payload},
        ],
        "temperature": 0.2,
        "max_tokens": 1800,
    }).encode("utf-8")

    req = urllib.request.Request(
        DEEPSEEK_URL,
        data=body,
        headers={
            "Authorization": "Bearer " + DEEPSEEK_KEY,
            "Content-Type": "application/json",
        },
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=60) as resp:
            data = json.loads(resp.read().decode("utf-8"))
    except urllib.error.HTTPError as e:
        sys.stderr.write("[WARN] DeepSeek HTTP %s: %s\n" % (e.code, e.reason))
        return None
    except Exception as e:
        sys.stderr.write("[WARN] DeepSeek request failed: %s\n" % e)
        return None

    try:
        content = data["choices"][0]["message"]["content"]
    except (KeyError, IndexError):
        sys.stderr.write("[WARN] DeepSeek unexpected response shape\n")
        return None

    # Extract the JSON array robustly (model sometimes wraps in ```json ... ```).
    start = content.find("[")
    end = content.rfind("]")
    if start == -1 or end == -1 or end < start:
        sys.stderr.write("[WARN] no JSON array in DeepSeek content\n")
        return None
    try:
        arr = json.loads(content[start:end + 1])
    except json.JSONDecodeError as e:
        sys.stderr.write("[WARN] DeepSeek JSON parse failed: %s\n" % e)
        return None

    out = {}
    for item in arr:
        if isinstance(item, dict) and "w" in item and "zh" in item:
            out[str(item["w"]).strip()] = str(item["zh"]).strip()
    return out


def translate_all(all_words, do_translate):
    """
    all_words: dict level_code -> list of word dicts.
    Returns dict word_lower -> zh (global cache keyed by word_lower + en_def hash).
    """
    cache = load_cache()

    # Build the set of (word, en_def) pairs that need translation.
    needed = []
    for code, rows in all_words.items():
        for row in rows:
            key = cache_key(row["word"], row["en_def"])
            if key not in cache:
                needed.append((row["word"], row["en_def"], key))

    total_needed = len(needed)
    if total_needed == 0:
        sys.stdout.write("[i] all translations cached (%d keys)\n" % len(cache))
        return cache

    if not do_translate:
        sys.stdout.write(
            "[i] %d translations missing; --translate not set -> skipping API calls\n"
            % total_needed
        )
        return cache

    sys.stdout.write("[i] translating %d definitions via DeepSeek...\n" % total_needed)
    done = 0
    # De-dup by (word, en_def) so we never call twice for identical pairs.
    unique_pairs = {}
    for word, en_def, key in needed:
        unique_pairs[key] = (word, en_def)
    pairs = list(unique_pairs.values())

    for i in range(0, len(pairs), BATCH_SIZE):
        batch = pairs[i:i + BATCH_SIZE]
        result = call_deepseek(batch)
        if result is None:
            sys.stderr.write("[WARN] batch %d-%d failed, will skip (cache keeps prior)\n"
                             % (i, i + len(batch)))
        else:
            for word, en_def in batch:
                zh = result.get(word)
                if not zh:
                    # try a case-insensitive match against returned keys
                    for k, v in result.items():
                        if k.lower() == word.lower():
                            zh = v
                            break
                if zh:
                    cache[cache_key(word, en_def)] = zh
                    done += 1
            save_cache(cache)
        if i + BATCH_SIZE < len(pairs):
            time.sleep(SLEEP_BETWEEN_BATCHES)
        if (i // BATCH_SIZE) % 5 == 0:
            sys.stdout.write("    progress %d/%d\n" % (i + len(batch), len(pairs)))
            sys.stdout.flush()

    sys.stdout.write("[i] translated %d/%d definitions\n" % (done, len(pairs)))
    return cache


def cache_key(word, en_def):
    return word.lower() + "||" + en_def.lower()


# ---------------------------------------------------------------------------
# DB writes (INSERT-only, idempotent)
# ---------------------------------------------------------------------------

def ensure_level(cur, code, name_zh, name_en, sort_order, target):
    """INSERT IGNORE into level (PK = code)."""
    cur.execute(
        "INSERT IGNORE INTO level (code, name_zh, name_en, sort_order, target_word_count) "
        "VALUES (%s, %s, %s, %s, %s)",
        (code, name_zh, name_en, sort_order, target),
    )


def insert_words(cur, level_code, rows, cache):
    """Idempotently insert words; return dict word_lower -> word_id.

    NOTE: word_bank.uk_level_word is (level_code, word_lower, deleted_at), and since
    deleted_at is NULL for live rows, MySQL treats them as DISTINCT (NULL != NULL) so
    INSERT IGNORE does NOT enforce uniqueness. We must SELECT-for-existence first.
    """
    inserted = 0
    word_ids = {}
    for row in rows:
        wl = row["word"].lower()
        if wl in word_ids:
            continue
        # Check existence explicitly (live = deleted_at IS NULL).
        cur.execute(
            "SELECT id FROM word_bank "
            "WHERE level_code=%s AND word_lower=%s AND deleted_at IS NULL LIMIT 1",
            (level_code, wl),
        )
        existing = cur.fetchone()
        if existing:
            word_ids[wl] = existing[0]
            continue
        en_def = row["en_def"] or row["word"]
        zh = cache.get(cache_key(row["word"], en_def)) or ""
        cur.execute(
            "INSERT INTO word_bank "
            "(level_code, word, word_lower, ipa_uk, en_definition, zh_definition, "
            " example_en, pos, audit_status) "
            "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, 1)",
            (level_code, row["word"], wl, row["ipa"], en_def, zh,
             row["example"], row["pos"]),
        )
        word_ids[wl] = cur.lastrowid
        inserted += 1
    sys.stdout.write("    word_bank %s: +%d new / %d total\n"
                     % (level_code, inserted, len(word_ids)))
    return word_ids


def ensure_list(cur, name_zh, level_code):
    """Idempotent by name. Returns list id."""
    cur.execute("SELECT id FROM word_list WHERE name=%s LIMIT 1", (name_zh,))
    row = cur.fetchone()
    if row:
        return row[0]
    cur.execute(
        "INSERT INTO word_list "
        "(owner_user_id, name, source_type, origin_level_code, cover_emoji, sort_order) "
        "VALUES (NULL, %s, 'builtin', %s, %s, 0)",
        (name_zh, level_code, "\U0001F4C8"),  # 📘
    )
    return cur.lastrowid


def insert_items(cur, list_id, rows, word_ids):
    """INSERT IGNORE word_list_item by (list_id, word_id)."""
    inserted = 0
    # preload existing (word_id) for this list
    cur.execute("SELECT word_id FROM word_list_item WHERE list_id=%s", (list_id,))
    existing = {r[0] for r in cur.fetchall()}
    for order, row in enumerate(rows, start=1):
        wid = word_ids.get(row["word"].lower())
        if not wid or wid in existing:
            continue
        cur.execute(
            "INSERT INTO word_list_item (list_id, word_id, unit_no, page, sort_order) "
            "VALUES (%s, %s, %s, %s, %s)",
            (list_id, wid, row["unit_no"], row["page"], order),
        )
        existing.add(wid)
        inserted += 1
    # update word_count to actual
    cur.execute(
        "UPDATE word_list SET word_count=(SELECT COUNT(*) FROM word_list_item WHERE list_id=%s) "
        "WHERE id=%s",
        (list_id, list_id),
    )
    return inserted


def evict_redis_cache():
    import subprocess
    for cmd in (REDIS_PATTERN_DEL_CMD,):
        try:
            out = subprocess.run(
                cmd, capture_output=True, text=True, timeout=15,
            )
            sys.stdout.write("[i] redis evict: %s -> rc=%s out=%s\n"
                             % (" ".join(cmd[-3:]), out.returncode, out.stdout.strip()))
            return
        except Exception as e:
            sys.stderr.write("[WARN] redis evict failed: %s\n" % e)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--translate", action="store_true",
                        help="Allow DeepSeek API calls (costs ~RMB 0.1).")
    args = parser.parse_args()

    if args.translate and not DEEPSEEK_KEY:
        sys.exit("[FATAL] --translate 需要环境变量 DEEPSEEK_API_KEY")

    # Force UTF-8 output (Windows GBK terminal chokes on IPA/CJK, see CLAUDE.md #28).
    try:
        sys.stdout.reconfigure(encoding="utf-8")
        sys.stderr.reconfigure(encoding="utf-8")
    except Exception:
        pass

    # 1. Find files & parse.
    # Note: on case-insensitive filesystems (Windows/macOS) globbing both
    # *.xlsx and *.XLSX double-counts files, so scan the dir once and match
    # the extension case-insensitively ourselves.
    files = []
    for f in sorted(os.listdir(WORDLIST_DIR)):
        if not f.lower().endswith(".xlsx"):
            continue
        full = os.path.join(WORDLIST_DIR, f)
        match = match_level_for_file(full)
        if match:
            files.append((full, match))
    if len(files) != len(LEVEL_MAP):
        sys.stderr.write("[FATAL] expected %d wordlist files, found %d\n"
                         % (len(LEVEL_MAP), len(files)))
        for f, _ in files:
            sys.stderr.write("   found: %s\n" % f)
        sys.exit(1)

    all_rows = {}
    total = 0
    for f, (frag, code, zh, en, sort, target) in files:
        rows = parse_excel(f)
        all_rows[code] = rows
        total += len(rows)
        sys.stdout.write("[parse] %s -> %s: parsed %d rows\n" % (os.path.basename(f), code, len(rows)))
    sys.stdout.write("[parse] TOTAL %d rows across %d levels\n" % (total, len(files)))

    # 2. Translate.
    cache = translate_all(all_rows, args.translate)
    cached_with_zh = sum(1 for code, rows in all_rows.items()
                         for r in rows
                         if cache.get(cache_key(r["word"], r["en_def"])))
    sys.stdout.write("[translate] have zh for %d/%d rows\n" % (cached_with_zh, total))

    # 3. DB writes.
    conn = pymysql.connect(**DB_CONFIG)
    try:
        with conn.cursor() as cur:
            for f, (frag, code, zh, en, sort, target) in files:
                sys.stdout.write("\n=== %s (%s) ===\n" % (code, zh))
                ensure_level(cur, code, zh, en, sort, target)
                word_ids = insert_words(cur, code, all_rows[code], cache)
                list_id = ensure_list(cur, zh, code)
                n_items = insert_items(cur, list_id, all_rows[code], word_ids)
                sys.stdout.write("    word_list %s (id=%s): +%d items\n" % (zh, list_id, n_items))
        conn.commit()
        sys.stdout.write("\n[DB] committed\n")
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

    # 4. Evict Redis cache (non-fatal).
    evict_redis_cache()

    sys.stdout.write("\n[DONE] imported %d words across %d lists\n" % (total, len(files)))


if __name__ == "__main__":
    main()
